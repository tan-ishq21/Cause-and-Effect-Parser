import os
import json
import pandas as pd
from flask import Flask, request, jsonify, render_template, redirect, url_for
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

app = Flask(__name__)

# ============================================================
# CONFIG
# ============================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

OUTPUT_JSON_PATH = os.path.join(BASE_DIR, "cause_effect_data.json")

ALLOWED_EXTENSIONS = {"xlsx"}

ALLOWED_MARKERS = {"S", "S*", "N", "N*", "NP", "NP*"}
ALLOWED_VOTING = {"1OO1", "1OO2", "2OO2", "1OO3", "2OO3", "3OO3", "2OO4", "3OO4"}

# Excel path will be set after upload
EXCEL_PATH = None

master_data = {}


# ============================================================
# UPLOAD HELPERS
# ============================================================
def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


# ============================================================
# EXCEL READ WITH MERGED CELLS FILLED
# ============================================================
def read_excel_sheet_with_merged(excel_path, sheet_name):
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]

    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))

    # Fill merged cells
    for merged_range in ws.merged_cells.ranges:
        min_row = merged_range.min_row - 1
        min_col = merged_range.min_col - 1
        max_row = merged_range.max_row - 1
        max_col = merged_range.max_col - 1

        top_left_value = data[min_row][min_col]

        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                data[r][c] = top_left_value

    return pd.DataFrame(data)


# ============================================================
# NORMALIZATION
# ============================================================
def normalize_text(val) -> str:
    if val is None:
        return ""
    text = str(val)
    text = text.replace("\n", " ").replace("\r", " ").replace("\t", " ")
    text = " ".join(text.split())
    return text.strip().upper()


def normalize_multiline_text(val) -> str:
    if val is None:
        return ""

    text = str(val)

    # keep line breaks, only clean spaces/tabs
    text = text.replace("\t", " ")
    text = text.replace("\r\n", "\n").replace("\r", "\n")

    # remove extra spaces inside each line
    lines = [(" ".join(line.split())).strip() for line in text.split("\n")]

    # remove empty lines
    lines = [line for line in lines if line]

    return "\n".join(lines).strip()


def normalize_cell(val) -> str:
    if pd.isna(val):
        return ""
    return normalize_text(val)


def find_cell(df, targets):
    targets_normalized = [normalize_text(t) for t in targets]

    for r in range(df.shape[0]):
        for c in range(df.shape[1]):
            cell = normalize_cell(df.iat[r, c])
            if cell in targets_normalized:
                return r, c
    return None, None


# ============================================================
# EFFECT REGION DETECTION
# ============================================================
def detect_effect_region(df):
    r_desc, c_desc = find_cell(df, ["Effect Description"])
    if r_desc is None:
        return None

    # Check if "Value" exists just below Effect Description
    possible_value_row = r_desc + 2
    value_cell = normalize_cell(df.iat[possible_value_row, c_desc])

    if value_cell == "VALUE":
        # OPERATIONAL SHEET FORMAT
        return {
            "has_identifier": False,
            "effect_identifier_row": None,
            "effect_description_row": r_desc,
            "value_row": r_desc + 2,
            "output_tag_row": r_desc + 3,
            "action_row": r_desc + 4,
            "matrix_start_row": r_desc + 5,
            "effects_start_col": c_desc + 1
        }

    # SIS FORMAT (NO VALUE ROW)
    return {
        "has_identifier": False,
        "effect_identifier_row": None,
        "effect_description_row": r_desc,
        "value_row": None,
        "output_tag_row": r_desc + 2,
        "action_row": r_desc + 3,
        "matrix_start_row": r_desc + 4,
        "effects_start_col": c_desc + 1
    }


def detect_effect_columns(df, effect_description_row, start_col):
    blank_count = 0
    end_col = start_col

    for c in range(start_col, df.shape[1]):
        header = normalize_cell(df.iat[effect_description_row, c])

        if header == "":
            blank_count += 1
            if blank_count >= 3:
                break
        else:
            blank_count = 0
            end_col = c

    return start_col, end_col


# ============================================================
# CAUSE HEADER DETECTION
# ============================================================
def detect_cause_columns(df):
    header_map = {
        "cause_identifier": ["Cause Identifier", "Cause ID", "Cause"],
        "input_tag": ["Input Tag", "InputTag", "Tag"],
        "signal": ["Signal"],
        "warn": ["Warn", "Warning"],
        "alarm": ["Alarm"],
        "safety_limit": ["Safety Limit", "SafetyLimit", "Switch/Limit"],
        "hyst": ["Hyst.", "Hysteresis"],
        "unit": ["Unit"],
        "delay": ["Delay"],
        "func1": ["Func 1", "Func1"],
        "func2": ["Func 2", "Func2"],
        "func3": ["Func 3", "Func3"],
        "cause_description": ["Cause Description", "Description"],
        "comment": ["Comment", "Remarks", "Remark"],
        "important_comment": ["Comments, Causes / Kommentare, Ursachen"],
        "w_dc": ["W_DC", "W DC", "W-DC"],
        "a_dc": ["A_DC", "A DC", "A-DC"],
        "a_dg": ["A_DG", "A DG", "A-DG"],
    }

    header_map_norm = {k: [normalize_text(x) for x in v] for k, v in header_map.items()}

    found_cols = {}
    header_row_found = None

    for r in range(df.shape[0]):
        row_values = [normalize_cell(x) for x in df.iloc[r].values]
        for c, val in enumerate(row_values):
            if val in header_map_norm["input_tag"]:
                found_cols["input_tag"] = c
                header_row_found = r
                break
        if header_row_found is not None:
            break

    if header_row_found is None:
        return None

    start_scan = max(0, header_row_found - 3)
    end_scan = min(df.shape[0], header_row_found + 8)

    for r in range(start_scan, end_scan):
        row_values = [normalize_cell(x) for x in df.iloc[r].values]
        for c, val in enumerate(row_values):
            for key, variants in header_map_norm.items():
                if val in variants:
                    found_cols[key] = c

    found_cols["_header_row"] = header_row_found
    return found_cols


# ============================================================
# MESSAGE ONLY SHEET EXTRACTION (FIXED)
# ============================================================
def extract_message_only_sheet(sheet_name, ws, df):
    """
    Extracts Ely System - Message Only sheet.
    Rule (same as SIS/Operational style):
    - If Func2 voting logic is merged vertically AND != 1OO1 -> show entire merged block.
    - If not merged OR voting == 1OO1 -> show only that row.
    """

    header_map = {
        "input_tag": ["Input Tag", "InputTag", "Tag"],
        "signal": ["Signal"],
        "warn": ["Warn", "Warning"],
        "alarm": ["Alarm"],
        "hyst": ["Hyst.", "Hysteresis"],
        "unit": ["Unit"],
        "w_dc": ["W_DC", "W DC", "W-DC"],
        "func1": ["Func 1", "Func1"],
        "func2": ["Func 2", "Func2"],
        "cause_description": ["Cause Description", "Description"],
        "comment": ["Comment", "Remarks", "Remark"],
    }

    header_map_norm = {k: [normalize_text(x) for x in v] for k, v in header_map.items()}

    found_cols = {}
    header_row_found = None

    # find header row by detecting Input Tag
    for r in range(df.shape[0]):
        row_values = [normalize_cell(x) for x in df.iloc[r].values]
        for c, val in enumerate(row_values):
            if val in header_map_norm["input_tag"]:
                found_cols["input_tag"] = c
                header_row_found = r
                break
        if header_row_found is not None:
            break

    if header_row_found is None:
        return []

    # scan around header row to find other columns
    start_scan = max(0, header_row_found - 3)
    end_scan = min(df.shape[0], header_row_found + 8)

    for r in range(start_scan, end_scan):
        row_values = [normalize_cell(x) for x in df.iloc[r].values]
        for c, val in enumerate(row_values):
            for key, variants in header_map_norm.items():
                if val in variants:
                    found_cols[key] = c

    if "input_tag" not in found_cols:
        return []

    merged_ranges = list(ws.merged_cells.ranges)

    def get_row_merge_id(row_0):
        excel_row = row_0 + 1

        for rng in merged_ranges:
            if rng.min_row <= excel_row <= rng.max_row:
                # must be vertical merge
                if rng.max_row > rng.min_row:
                    return f"{rng.min_row}:{rng.min_col}-{rng.max_row}:{rng.max_col}"

        return None

    # Voting logic is normally in FUNC2 (same style as SIS/Operational voting col)
    logic_col = found_cols.get("func2", None)
    if logic_col is None:
        logic_col = found_cols.get("func1", None)

    if logic_col is None:
        return []

    records = []
    visited_merge_ids = set()

    empty_count = 0

    for r in range(header_row_found + 1, df.shape[0]):

        input_tag = normalize_cell(df.iat[r, found_cols["input_tag"]])

        if input_tag == "":
            empty_count += 1
            if empty_count >= 15:
                break
            continue
        else:
            empty_count = 0

        logic_val = normalize_cell(df.iat[r, logic_col]).replace(" ", "")
        merge_id = get_row_merge_id(r)

        # SIS-style: if merged block and voting != 1OO1, show full block
        if merge_id is not None and logic_val != "1OO1":

            if merge_id in visited_merge_ids:
                continue

            visited_merge_ids.add(merge_id)

            left, right = merge_id.split("-")
            min_row, _ = left.split(":")
            max_row, _ = right.split(":")

            min_row = int(min_row) - 1
            max_row = int(max_row) - 1

            for rr in range(min_row, max_row + 1):

                row_input_tag = normalize_cell(df.iat[rr, found_cols["input_tag"]])
                if row_input_tag == "":
                    continue

                records.append({
                    "sheet": sheet_name,
                    "row": rr,
                    "input_tag": row_input_tag,
                    "signal": normalize_cell(df.iat[rr, found_cols["signal"]]) if "signal" in found_cols else "",
                    "warn": normalize_cell(df.iat[rr, found_cols["warn"]]) if "warn" in found_cols else "",
                    "alarm": normalize_cell(df.iat[rr, found_cols["alarm"]]) if "alarm" in found_cols else "",
                    "hyst": normalize_cell(df.iat[rr, found_cols["hyst"]]) if "hyst" in found_cols else "",
                    "unit": normalize_cell(df.iat[rr, found_cols["unit"]]) if "unit" in found_cols else "",
                    "w_dc": normalize_cell(df.iat[rr, found_cols["w_dc"]]) if "w_dc" in found_cols else "",
                    "func1": normalize_cell(df.iat[rr, found_cols["func1"]]) if "func1" in found_cols else "",
                    "func2": normalize_cell(df.iat[rr, found_cols["func2"]]) if "func2" in found_cols else "",
                    "cause_description": normalize_cell(df.iat[rr, found_cols["cause_description"]]) if "cause_description" in found_cols else "",
                    "comment": normalize_multiline_text(df.iat[rr, found_cols["comment"]]) if "comment" in found_cols else ""
                })

        else:
            # Not merged OR logic is 1OO1 -> only single row
            records.append({
                "sheet": sheet_name,
                "row": r,
                "input_tag": input_tag,
                "signal": normalize_cell(df.iat[r, found_cols["signal"]]) if "signal" in found_cols else "",
                "warn": normalize_cell(df.iat[r, found_cols["warn"]]) if "warn" in found_cols else "",
                "alarm": normalize_cell(df.iat[r, found_cols["alarm"]]) if "alarm" in found_cols else "",
                "hyst": normalize_cell(df.iat[r, found_cols["hyst"]]) if "hyst" in found_cols else "",
                "unit": normalize_cell(df.iat[r, found_cols["unit"]]) if "unit" in found_cols else "",
                "w_dc": normalize_cell(df.iat[r, found_cols["w_dc"]]) if "w_dc" in found_cols else "",
                "func1": normalize_cell(df.iat[r, found_cols["func1"]]) if "func1" in found_cols else "",
                "func2": normalize_cell(df.iat[r, found_cols["func2"]]) if "func2" in found_cols else "",
                "cause_description": normalize_cell(df.iat[r, found_cols["cause_description"]]) if "cause_description" in found_cols else "",
                "comment": normalize_multiline_text(df.iat[r, found_cols["comment"]]) if "comment" in found_cols else ""
            })

    return records


# ============================================================
# DYNAMIC LOGIC COLUMN DETECTION
# ============================================================
def detect_logic_columns(df, cause_cols, start_row):
    func_candidates = []
    for key in ["func1", "func2", "func3"]:
        if key in cause_cols:
            func_candidates.append((key, cause_cols[key]))

    if not func_candidates:
        return {"and_col": None, "voting_col": None}

    scan_end = min(df.shape[0], start_row + 200)

    and_counts = {col: 0 for _, col in func_candidates}
    voting_counts = {col: 0 for _, col in func_candidates}

    for r in range(start_row, scan_end):
        for _, col in func_candidates:
            val = normalize_cell(df.iat[r, col])
            val_clean = val.replace(" ", "")

            if val == "AND":
                and_counts[col] += 1
            if val_clean in ALLOWED_VOTING:
                voting_counts[col] += 1

    and_col = max(and_counts, key=and_counts.get) if max(and_counts.values()) > 0 else None
    voting_col = max(voting_counts, key=voting_counts.get) if max(voting_counts.values()) > 0 else None

    return {"and_col": and_col, "voting_col": voting_col}


# ============================================================
# SHEET EXTRACTION WITH SAFETY LOGIC BLOCKS
# ============================================================
def extract_sheet(sheet_name, ws, df):
    region = detect_effect_region(df)
    if not region:
        print(f"[SKIP] No Effect region in sheet: {sheet_name}")
        return {"records": [], "logic_blocks": []}

    effects_start_col, effects_end_col = detect_effect_columns(
        df,
        region["effect_description_row"],
        region["effects_start_col"]
    )

    effects_metadata = []
    for c in range(effects_start_col, effects_end_col + 1):

        effect_id = ""
        if region["has_identifier"]:
            effect_id = normalize_cell(df.iat[region["effect_identifier_row"], c])

        effect_desc = normalize_cell(df.iat[region["effect_description_row"], c])
        value = ""
        if region.get("value_row") is not None:
            value = normalize_cell(df.iat[region["value_row"], c])

        output_tag = normalize_cell(df.iat[region["output_tag_row"], c])
        action = normalize_cell(df.iat[region["action_row"], c])

        if effect_desc == "" and effect_id == "" and output_tag == "" and action == "":
            continue

        unique_effect_key = " | ".join([
            effect_id if effect_id else effect_desc,
            output_tag,
            action,
            f"COL_{c}"
        ])

        effects_metadata.append({
            "col": c,
            "effect_key": unique_effect_key,
            "effect_id": effect_id if effect_id else effect_desc,
            "effect_desc": effect_desc,
            "output_tag": output_tag,
            "action": action,
            "value": value
        })

    cause_cols = detect_cause_columns(df)
    if not cause_cols or "input_tag" not in cause_cols:
        print(f"[SKIP] No Input Tag column in sheet: {sheet_name}")
        return {"records": [], "logic_blocks": []}

    logic_cols = detect_logic_columns(df, cause_cols, region["matrix_start_row"])
    and_col = logic_cols["and_col"]
    voting_col = logic_cols["voting_col"]

    if and_col is None or voting_col is None:
        print(f"[WARN] Missing AND/Voting cols in sheet {sheet_name}")
        return {"records": [], "logic_blocks": []}

    merged_ranges = list(ws.merged_cells.ranges)

    def is_cell_strikethrough(row_0, col_0):
        row = row_0 + 1
        col = col_0 + 1

        for rng in merged_ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                top_left_cell = ws.cell(row=rng.min_row, column=rng.min_col)
                return top_left_cell.font is not None and top_left_cell.font.strike is True

        cell = ws.cell(row=row, column=col)
        return cell.font is not None and cell.font.strike is True

    def get_merged_range_id(row_0, col_0):
        row = row_0 + 1
        col = col_0 + 1

        for rng in merged_ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                return f"{rng.min_row}:{rng.min_col}-{rng.max_row}:{rng.max_col}"
        return None

    records = []
    logic_blocks = []

    current_block = None
    current_and_group = None

    prev_voting_merge_id = None
    prev_and_merge_id = None

    empty_count = 0

    for r in range(region["matrix_start_row"], df.shape[0]):

        input_tag = normalize_cell(df.iat[r, cause_cols["input_tag"]])

        input_tag_col = cause_cols["input_tag"]
        cause_id_col = cause_cols.get("cause_identifier", None)

        input_tag_strike = is_cell_strikethrough(r, input_tag_col)

        cause_id_strike = False
        if cause_id_col is not None:
            cause_id_strike = is_cell_strikethrough(r, cause_id_col)

        if input_tag_strike or cause_id_strike:
            continue

        if input_tag == "":
            empty_count += 1
            if empty_count >= 15:
                break
            continue
        else:
            empty_count = 0

        and_val = normalize_cell(df.iat[r, and_col])
        voting_val = normalize_cell(df.iat[r, voting_col]).replace(" ", "")

        cause_desc = normalize_cell(df.iat[r, cause_cols["cause_description"]]) if "cause_description" in cause_cols else ""
        comment = normalize_cell(df.iat[r, cause_cols["comment"]]) if "comment" in cause_cols else ""
        important_comment = normalize_multiline_text(df.iat[r, cause_cols["important_comment"]]) if "important_comment" in cause_cols else ""

        voting_merge_id = get_merged_range_id(r, voting_col)
        and_merge_id = get_merged_range_id(r, and_col)

        row_effects = []
        for eff in effects_metadata:
            marker = normalize_cell(df.iat[r, eff["col"]])

            if marker not in ALLOWED_MARKERS:
                continue

            if is_cell_strikethrough(r, eff["col"]):
                continue

            row_effects.append({
                "effect_key": eff["effect_key"],
                "effect_id": eff["effect_id"],
                "effect_desc": eff["effect_desc"],
                "output_tag": eff["output_tag"],
                "action": eff["action"],
                "value": eff.get("value", ""),
                "marker": marker
            })

        record = {
            "sheet": sheet_name,
            "row": r,
            "input_tag": input_tag,
            "cause_identifier": normalize_cell(df.iat[r, cause_cols["cause_identifier"]]) if "cause_identifier" in cause_cols else "",
            "signal": normalize_cell(df.iat[r, cause_cols["signal"]]) if "signal" in cause_cols else "",
            "warn": normalize_cell(df.iat[r, cause_cols["warn"]]) if "warn" in cause_cols else "",
            "w_dc": normalize_cell(df.iat[r, cause_cols["w_dc"]]) if "w_dc" in cause_cols else "",
            "a_dc": normalize_cell(df.iat[r, cause_cols["a_dc"]]) if "a_dc" in cause_cols else "",
            "a_dg": normalize_cell(df.iat[r, cause_cols["a_dg"]]) if "a_dg" in cause_cols else "",
            "safety_limit": normalize_cell(df.iat[r, cause_cols["safety_limit"]]) if "safety_limit" in cause_cols else "",
            "hyst": normalize_cell(df.iat[r, cause_cols["hyst"]]) if "hyst" in cause_cols else "",
            "unit": normalize_cell(df.iat[r, cause_cols["unit"]]) if "unit" in cause_cols else "",
            "delay": normalize_cell(df.iat[r, cause_cols["delay"]]) if "delay" in cause_cols else "",
            "cause_description": cause_desc,
            "comment": comment,
            "important_comment": important_comment,
            "and_logic": and_val,
            "voting_logic": voting_val,
            "effects": row_effects
        }

        records.append(record)

        start_new_block = False

        if voting_merge_id is not None:
            if voting_merge_id != prev_voting_merge_id:
                start_new_block = True
        else:
            start_new_block = True

        if start_new_block:
            if current_block is not None:
                logic_blocks.append(current_block)

            current_block = {
                "sheet": sheet_name,
                "block_id": f"{normalize_text(sheet_name)}_BLOCK_{len(logic_blocks)+1}",
                "voting_logic": voting_val,
                "cause_description": cause_desc,
                "comment": comment,
                "important_comment": important_comment,
                "has_and_logic": False,
                "and_groups": [],
                "effects": {}
            }

            current_and_group = None
            prev_and_merge_id = None

        if current_block is None:
            current_block = {
                "sheet": sheet_name,
                "block_id": f"{normalize_text(sheet_name)}_BLOCK_{len(logic_blocks)+1}",
                "voting_logic": voting_val,
                "cause_description": cause_desc,
                "comment": comment,
                "important_comment": important_comment,
                "has_and_logic": False,
                "and_groups": [],
                "effects": {}
            }

        if and_val in ["AND", "OR"]:
            current_block["has_and_logic"] = True

            if and_merge_id != prev_and_merge_id or and_merge_id is None:
                current_and_group = {
                    "group_id": f"{and_val}_{len(current_block['and_groups'])+1}",
                    "logic": and_val,
                    "tags": []
                }
                current_block["and_groups"].append(current_and_group)

        else:
            current_and_group = {
                "group_id": f"SINGLE_{len(current_block['and_groups'])+1}",
                "logic": "SINGLE",
                "tags": []
            }
            current_block["and_groups"].append(current_and_group)
            prev_and_merge_id = None

        if current_and_group is None:
            current_and_group = {
                "group_id": f"SINGLE_{len(current_block['and_groups'])+1}",
                "logic": "SINGLE",
                "tags": []
            }
            current_block["and_groups"].append(current_and_group)

        current_and_group["tags"].append({
            "input_tag": record["input_tag"],
            "cause_identifier": record["cause_identifier"],
            "signal": record["signal"],
            "warn": record["warn"],
            "limit": record["safety_limit"],
            "hyst": record["hyst"],
            "unit": record["unit"],
            "delay": record["delay"],
            "comment": record["comment"],
            "important_comment": record["important_comment"]
        })

        for eff in row_effects:
            key = eff["effect_key"]

            if key not in current_block["effects"]:
                current_block["effects"][key] = {
                    "effect_id": eff["effect_id"],
                    "effect_desc": eff["effect_desc"],
                    "output_tag": eff["output_tag"],
                    "action": eff["action"],
                    "value": eff.get("value", ""),
                    "markers": []
                }

            if eff["marker"] not in current_block["effects"][key]["markers"]:
                current_block["effects"][key]["markers"].append(eff["marker"])

        prev_voting_merge_id = voting_merge_id
        prev_and_merge_id = and_merge_id

    if current_block is not None:
        logic_blocks.append(current_block)

    return {"records": records, "logic_blocks": logic_blocks}


# ============================================================
# MASTER JSON BUILD
# ============================================================
def build_master_json(all_records, all_logic_blocks, message_only_records):
    index_by_input_tag = {}
    index_by_cause_identifier = {}
    index_by_logic_block_tag = {}

    for i, rec in enumerate(all_records):
        tag = normalize_text(rec.get("input_tag", ""))
        if tag:
            index_by_input_tag.setdefault(tag, []).append(i)

        sheet_name = normalize_text(rec.get("sheet", ""))

        if sheet_name == "ELY SYSTEM - SIS":
            cid = normalize_text(rec.get("cause_identifier", ""))
            if cid:
                index_by_cause_identifier.setdefault(cid, []).append(i)

    for bi, block in enumerate(all_logic_blocks):
        for grp in block.get("and_groups", []):
            for tag_obj in grp.get("tags", []):
                t = normalize_text(tag_obj.get("input_tag", ""))
                if t:
                    index_by_logic_block_tag.setdefault(t, []).append(bi)

    return {
        "total_records": len(all_records),
        "total_logic_blocks": len(all_logic_blocks),
        "records": all_records,
        "logic_blocks": all_logic_blocks,
        "message_only_records": message_only_records,
        "index_by_input_tag": index_by_input_tag,
        "index_by_cause_identifier": index_by_cause_identifier,
        "index_by_logic_block_tag": index_by_logic_block_tag
    }


def save_master_json(master_json):
    with open(OUTPUT_JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(master_json, f, indent=4, ensure_ascii=False)

    print(f"[INFO] JSON saved: {OUTPUT_JSON_PATH}")


def create_json_from_excel():
    global EXCEL_PATH

    if not EXCEL_PATH or not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"Excel not found: {EXCEL_PATH}")

    wb = load_workbook(EXCEL_PATH, data_only=True)

    all_records = []
    all_logic_blocks = []
    message_only_records = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        df = read_excel_sheet_with_merged(EXCEL_PATH, sheet)

        if normalize_text(sheet) == "ELY SYSTEM - MESSAGE ONLY":
            extracted_msgs = extract_message_only_sheet(sheet, ws, df)
            message_only_records.extend(extracted_msgs)
            continue

        extracted = extract_sheet(sheet, ws, df)
        all_records.extend(extracted["records"])
        all_logic_blocks.extend(extracted["logic_blocks"])

    master_json = build_master_json(all_records, all_logic_blocks, message_only_records)
    save_master_json(master_json)


def load_master_json():
    global master_data

    if not os.path.exists(OUTPUT_JSON_PATH):
        print("[INFO] JSON not found. Creating JSON from Excel...")
        create_json_from_excel()

    with open(OUTPUT_JSON_PATH, "r", encoding="utf-8") as f:
        master_data = json.load(f)

    print(f"[INFO] Loaded records: {master_data.get('total_records', 0)}")
    print(f"[INFO] Loaded logic blocks: {master_data.get('total_logic_blocks', 0)}")
    print(f"[INFO] Loaded message only rows: {len(master_data.get('message_only_records', []))}")


# ============================================================
# SEARCH
# ============================================================
def search_logic_blocks(query):
    q = normalize_text(query)

    block_indexes = set()
    for tag, idx_list in master_data.get("index_by_logic_block_tag", {}).items():
        if q in tag:
            block_indexes.update(idx_list)

    grouped = {}
    for bi in block_indexes:
        block = master_data["logic_blocks"][bi]
        grouped.setdefault(block["sheet"], []).append(block)

    return grouped


def search_records(query):
    q = normalize_text(query)

    record_indexes = set()

    for key, idx_list in master_data.get("index_by_input_tag", {}).items():
        if q in key:
            record_indexes.update(idx_list)

    for key, idx_list in master_data.get("index_by_cause_identifier", {}).items():
        if q in key:
            record_indexes.update(idx_list)

    grouped = {}
    for idx in record_indexes:
        rec = master_data["records"][idx]
        grouped.setdefault(rec["sheet"], []).append(rec)

    for sheet in grouped:
        grouped[sheet].sort(key=lambda x: x.get("input_tag", ""))

    return grouped


def search_message_only(query):
    q = normalize_text(query)
    results = []

    for rec in master_data.get("message_only_records", []):
        tag = normalize_text(rec.get("input_tag", ""))
        if q in tag:
            results.append(rec)

    results.sort(key=lambda x: x.get("input_tag", ""))
    return results


# ============================================================
# FLASK ROUTES
# ============================================================

@app.route("/", methods=["GET", "POST"])
def landing():
    """
    Landing page where user uploads excel file.
    """
    global EXCEL_PATH

    if request.method == "POST":

        if "file" not in request.files:
            return render_template("landing.html", error="No file part found!")

        file = request.files["file"]

        if file.filename == "":
            return render_template("landing.html", error="No file selected!")

        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            save_path = os.path.join(UPLOAD_FOLDER, filename)
            file.save(save_path)

            EXCEL_PATH = save_path

            # rebuild JSON and reload
            create_json_from_excel()
            load_master_json()

            return redirect(url_for("home"))

        return render_template("landing.html", error="Invalid file type! Upload only .xlsx file.")

    return render_template("landing.html")


@app.route("/main")
def home():
    return render_template("index.html")


@app.route("/search", methods=["GET"])
def search_api():
    query = request.args.get("q", "").strip()
    if not query:
        return jsonify({"error": "Empty search query"}), 400

    record_results = search_records(query)
    logic_results = search_logic_blocks(query)

    message_results = search_message_only(query)

    return jsonify({
        "query": query,
        "record_count": sum(len(v) for v in record_results.values()),
        "logic_block_count": sum(len(v) for v in logic_results.values()),
        "message_only_count": len(message_results),
        "records": record_results,
        "logic_blocks": logic_results,
        "message_only": message_results
    })


@app.route("/reload", methods=["GET"])
def reload_data():
    create_json_from_excel()
    load_master_json()
    return jsonify({"status": "JSON rebuilt and reloaded successfully"})


@app.route("/debug_keys", methods=["GET"])
def debug_keys():
    return jsonify({
        "sample_input_tag_keys": list(master_data.get("index_by_input_tag", {}).keys())[:20],
        "sample_logic_block_keys": list(master_data.get("index_by_logic_block_tag", {}).keys())[:20],
        "total_blocks": master_data.get("total_logic_blocks", 0)
    })


# ============================================================
# MAIN
# ============================================================
if __name__ == "__main__":
    # DO NOT load JSON at startup because excel will be uploaded first
    app.run(host="0.0.0.0", port=5000, debug=True)